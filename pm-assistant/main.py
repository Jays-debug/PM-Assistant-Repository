import datetime
import csv
import io
import os
import json
import hmac
import hashlib
import base64
import requests
import shutil
import zipfile
import platform
import uuid
from pathlib import Path
from openpyxl import Workbook, load_workbook
from fastapi import FastAPI, Depends, HTTPException, Query, status, UploadFile, File, Request, Header
from fastapi.staticfiles import StaticFiles
from fastapi.responses import FileResponse, RedirectResponse, StreamingResponse
from fastapi.middleware.cors import CORSMiddleware
from sqlalchemy.orm import Session
from typing import List, Optional
from pydantic import BaseModel, Field

import database
from database import get_db, PMPlan, Location, Setting, init_db, VehicleMaster, PMHistory, NotificationLog, ImportLog, UserMaster, LineTarget, LineWebhookEvent, PMTaskState, WeeklyCampaign, WeeklyCampaignItem, import_data_car_to_vehicle_master
from notifier import trigger_daily_notification, send_line_message, get_alert_summary_text, send_line_message_debug

from apscheduler.schedulers.background import BackgroundScheduler
from apscheduler.triggers.cron import CronTrigger

app = FastAPI(title="PM Vehicle Tracking System")

LOG_DIR = Path(__file__).resolve().parent / "logs"
LOG_DIR.mkdir(exist_ok=True)

def write_app_log(filename: str, message: str):
    try:
        with open(LOG_DIR / filename, "a", encoding="utf-8") as f:
            f.write(f"{datetime.datetime.now().isoformat(timespec='seconds')} {message}\n")
    except Exception:
        pass


# Enable CORS (for development convenience)
app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

# Initialize database tables & default data
init_db()



def load_car_master_map():
    # v1.3: Vehicle master is stored in SQLite table vehicle_master.
    car_map = {}
    db = database.SessionLocal()
    try:
        vehicles = db.query(VehicleMaster).all()
        for v in vehicles:
            car_map[v.vehicle_no] = {
                "car_no": v.vehicle_no,
                "code": v.vehicle_code or "",
                "type": v.transport_type or "",
                "model": v.vehicle_model or "",
                "fleet": v.fleet or "",
                "vehicle_type": v.vehicle_type or ""
            }
        if car_map:
            return car_map
    except Exception as e:
        print("Error loading vehicle_master:", e)
    finally:
        db.close()

    # Fallback for first-run or damaged database: read Data Car.csv.
    csv_path = os.path.join(os.path.dirname(__file__), "Data Car.csv")
    car_map = {}
    if not os.path.exists(csv_path):
        return car_map
    try:
        with open(csv_path, mode="r", encoding="cp874", errors="ignore") as f:
            reader = csv.reader(f)
            next(reader, None)
            for row in reader:
                if len(row) >= 5:
                    car_no = row[4].strip()
                    if not car_no:
                        continue
                    car_map[car_no] = {
                        "car_no": car_no,
                        "code": row[0].strip(),
                        "type": row[1].strip(),
                        "model": row[2].strip(),
                        "fleet": row[3].strip(),
                        "vehicle_type": row[5].strip() if len(row) >= 6 else ""
                    }
    except Exception as e:
        print("Error loading car master map:", e)
    return car_map

def extract_car_no(vehicle_text: str) -> str:
    return (vehicle_text or "").split(" | ")[0].strip()

def enrich_plan_vehicle_fields(db_plan):
    car_map = load_car_master_map()
    car_no = extract_car_no(db_plan.vehicle_no)
    car = car_map.get(car_no)
    if car:
        db_plan.vehicle_code = car.get("code")
        db_plan.vehicle_model = car.get("model")
        db_plan.transport_type = car.get("type")
        db_plan.fleet = car.get("fleet")
        db_plan.vehicle_type = car.get("vehicle_type")

# Pydantic Schemas
class PMPlanSchema(BaseModel):
    id: int
    vehicle_no: str
    job_title: str
    description: Optional[str] = None
    planned_date: datetime.date
    actual_date: Optional[datetime.date] = None
    deadline_date: datetime.date
    location: str
    vehicle_code: Optional[str] = None
    vehicle_model: Optional[str] = None
    transport_type: Optional[str] = None
    fleet: Optional[str] = None
    vehicle_type: Optional[str] = None
    status: str
    created_at: datetime.datetime

    class Config:
        from_attributes = True

class PMPlanCreate(BaseModel):
    vehicle_no: str = Field(..., min_length=1)
    job_title: str = Field(..., min_length=1)
    description: Optional[str] = None
    planned_date: datetime.date
    actual_date: Optional[datetime.date] = None
    deadline_date: datetime.date
    location: str = Field(..., min_length=1)
    vehicle_code: Optional[str] = None
    vehicle_model: Optional[str] = None
    transport_type: Optional[str] = None
    fleet: Optional[str] = None
    vehicle_type: Optional[str] = None
    status: str = "Planned"

class LocationSchema(BaseModel):
    id: int
    name: str
    province: Optional[str] = None
    district: Optional[str] = None
    service_type: str = "external"
    address: Optional[str] = None
    note: Optional[str] = None

    class Config:
        from_attributes = True

class LocationCreate(BaseModel):
    name: str = Field(..., min_length=1)
    province: Optional[str] = None
    district: Optional[str] = None
    service_type: str = "external"
    address: Optional[str] = None
    note: Optional[str] = None

class BulkDeleteRequest(BaseModel):
    ids: List[int]

class SettingsSchema(BaseModel):
    line_channel_access_token: str
    line_target_id: str
    line_channel_secret: Optional[str] = None
    line_webhook_enabled: Optional[bool] = True
    alert_days_before: int
    scheduler_time: str
    scheduler_enabled: bool
    notification_template: Optional[str] = None
    debug_mode: Optional[bool] = False
    report_morning_time: Optional[str] = "07:30"
    report_noon_time: Optional[str] = "12:00"
    report_evening_time: Optional[str] = "17:00"
    report_morning_enabled: Optional[bool] = True
    report_noon_enabled: Optional[bool] = True
    report_evening_enabled: Optional[bool] = True

class LineTargetSchema(BaseModel):
    id: int
    source_type: str
    source_id: str
    display_name: Optional[str] = None
    is_active: bool
    last_event_at: datetime.datetime
    class Config:
        from_attributes = True

class UseLineTargetPayload(BaseModel):
    target_id: str

class WebhookTestPayload(BaseModel):
    text: Optional[str] = None

class ImportPreviewRequest(BaseModel):
    token: str

class ImportConfirmRequest(BaseModel):
    token: str

class PMHistorySchema(BaseModel):
    id: int
    pm_plan_id: int
    action: str
    changed_by: str
    changed_at: datetime.datetime
    old_data: Optional[str] = None
    new_data: Optional[str] = None
    note: Optional[str] = None
    class Config:
        from_attributes = True

class NotificationLogSchema(BaseModel):
    id: int
    channel: str
    target_id: Optional[str] = None
    message: Optional[str] = None
    status: str
    response: Optional[str] = None
    sent_at: datetime.datetime
    class Config:
        from_attributes = True

class ImportLogSchema(BaseModel):
    id: int
    import_type: str
    filename: Optional[str] = None
    status: str
    total_rows: int
    valid_rows: int
    error_rows: int
    imported_rows: int
    errors: Optional[str] = None
    created_by: str
    created_at: datetime.datetime
    class Config:
        from_attributes = True

class VehicleMasterSchema(BaseModel):
    id: int
    vehicle_no: str
    vehicle_code: Optional[str] = None
    transport_type: Optional[str] = None
    vehicle_model: Optional[str] = None
    fleet: Optional[str] = None
    vehicle_type: Optional[str] = None
    class Config:
        from_attributes = True

# Background Scheduler setup
scheduler = BackgroundScheduler(timezone="Asia/Bangkok")
scheduler.start()

def daily_notification_job():
    print("[Scheduler] Running daily scheduled notification job...")
    db_session = database.SessionLocal()
    try:
        success = trigger_daily_notification(db_session)
        print(f"[Scheduler] Daily notification job completed. Success: {success}")
    except Exception as e:
        print(f"[Scheduler] Error running daily notification job: {e}")
    finally:
        db_session.close()

def update_scheduler_job(db: Session):
    """
    Updates the background scheduler job based on current database settings.
    """
    job_id = "daily_pm_report"
    if scheduler.get_job(job_id):
        scheduler.remove_job(job_id)
        print("[Scheduler] Existing job removed.")

    token_setting = db.query(Setting).filter(Setting.key == "line_channel_access_token").first()
    target_setting = db.query(Setting).filter(Setting.key == "line_target_id").first()
    time_setting = db.query(Setting).filter(Setting.key == "scheduler_time").first()
    enabled_setting = db.query(Setting).filter(Setting.key == "scheduler_enabled").first()

    token = token_setting.value if token_setting else ""
    target_id = target_setting.value if target_setting else ""
    sched_time = time_setting.value if time_setting else "08:00"
    enabled = (enabled_setting.value.lower() == "true") if enabled_setting else False

    if enabled and token and target_id:
        try:
            hour, minute = map(int, sched_time.split(":"))
            trigger = CronTrigger(hour=hour, minute=minute, timezone="Asia/Bangkok")
            scheduler.add_job(
                daily_notification_job,
                trigger=trigger,
                id=job_id,
                replace_existing=True
            )
            print(f"[Scheduler] Job scheduled daily at {sched_time} Bangkok time.")
        except Exception as e:
            print(f"[Scheduler] Error setting up scheduler: {e}")
    else:
        print("[Scheduler] Job is disabled or Line details are not fully configured.")

# Initialize scheduler on startup
db_session = database.SessionLocal()
try:
    update_scheduler_job(db_session)
finally:
    db_session.close()



def _plan_to_dict(plan):
    return {
        "id": plan.id, "vehicle_no": plan.vehicle_no, "job_title": plan.job_title,
        "description": plan.description, "planned_date": str(plan.planned_date) if plan.planned_date else None,
        "actual_date": str(plan.actual_date) if plan.actual_date else None,
        "deadline_date": str(plan.deadline_date) if plan.deadline_date else None, "location": plan.location,
        "vehicle_code": plan.vehicle_code, "vehicle_model": plan.vehicle_model, "transport_type": plan.transport_type,
        "fleet": plan.fleet, "vehicle_type": plan.vehicle_type, "status": plan.status
    }

def add_pm_history(db, plan, action, old_data=None, note=None):
    try:
        db.add(PMHistory(
            pm_plan_id=plan.id if plan and plan.id else 0,
            action=action,
            changed_by="system",
            old_data=json.dumps(old_data, ensure_ascii=False) if old_data is not None else None,
            new_data=json.dumps(_plan_to_dict(plan), ensure_ascii=False) if plan is not None else None,
            note=note
        ))
    except Exception as e:
        print("History log warning:", e)

def add_import_log(db, import_type, filename, status, total=0, valid=0, error=0, imported=0, errors=None):
    log = ImportLog(import_type=import_type, filename=filename, status=status, total_rows=total, valid_rows=valid, error_rows=error, imported_rows=imported, errors=json.dumps(errors or [], ensure_ascii=False))
    db.add(log)
    return log

# API Routes

# 1. PM Plans CRUD
@app.get("/api/plans", response_model=List[PMPlanSchema])
def get_plans(
    status_filter: Optional[str] = Query(None, alias="status"),
    search: Optional[str] = None,
    next_day_only: bool = False,
    db: Session = Depends(get_db)
):
    query = db.query(PMPlan)
    
    if next_day_only:
        tomorrow = datetime.date.today() + datetime.timedelta(days=1)
        query = query.filter(PMPlan.planned_date == tomorrow)
    
    if status_filter:
        query = query.filter(PMPlan.status == status_filter)
        
    if search:
        search_filter = f"%{search}%"
        query = query.filter(
            (PMPlan.vehicle_no.like(search_filter)) |
            (PMPlan.job_title.like(search_filter)) |
            (PMPlan.description.like(search_filter)) |
            (PMPlan.location.like(search_filter)) |
            (PMPlan.vehicle_model.like(search_filter)) |
            (PMPlan.transport_type.like(search_filter)) |
            (PMPlan.fleet.like(search_filter)) |
            (PMPlan.vehicle_type.like(search_filter))
        )
        
    return query.order_by(PMPlan.planned_date.asc()).all()

@app.post("/api/plans", response_model=PMPlanSchema)
def create_plan(plan_in: PMPlanCreate, db: Session = Depends(get_db)):
    db_plan = PMPlan(
        vehicle_no=plan_in.vehicle_no,
        job_title=plan_in.job_title,
        description=plan_in.description,
        planned_date=plan_in.planned_date,
        actual_date=plan_in.actual_date,
        deadline_date=plan_in.deadline_date,
        location=plan_in.location,
        vehicle_code=plan_in.vehicle_code,
        vehicle_model=plan_in.vehicle_model,
        transport_type=plan_in.transport_type,
        fleet=plan_in.fleet,
        vehicle_type=plan_in.vehicle_type,
        status=plan_in.status
    )
    
    enrich_plan_vehicle_fields(db_plan)
    if db_plan.status != "Completed" and db_plan.status != "Cancelled" and db_plan.deadline_date < datetime.date.today():
        db_plan.status = "Overdue"
        
    db.add(db_plan)
    db.flush()
    add_pm_history(db, db_plan, "create", note="สร้างแผนงาน PM")
    db.commit()
    db.refresh(db_plan)
    return db_plan

@app.put("/api/plans/{plan_id}", response_model=PMPlanSchema)
def update_plan(plan_id: int, plan_in: PMPlanCreate, db: Session = Depends(get_db)):
    db_plan = db.query(PMPlan).filter(PMPlan.id == plan_id).first()
    if not db_plan:
        raise HTTPException(status_code=404, detail="Plan not found")

    old_data = _plan_to_dict(db_plan)
        
    db_plan.vehicle_no = plan_in.vehicle_no
    db_plan.job_title = plan_in.job_title
    db_plan.description = plan_in.description
    db_plan.planned_date = plan_in.planned_date
    db_plan.actual_date = plan_in.actual_date
    db_plan.deadline_date = plan_in.deadline_date
    db_plan.location = plan_in.location
    db_plan.vehicle_code = plan_in.vehicle_code
    db_plan.vehicle_model = plan_in.vehicle_model
    db_plan.transport_type = plan_in.transport_type
    db_plan.fleet = plan_in.fleet
    db_plan.vehicle_type = plan_in.vehicle_type
    db_plan.status = plan_in.status
    enrich_plan_vehicle_fields(db_plan)
    
    if db_plan.status != "Completed" and db_plan.status != "Cancelled" and db_plan.deadline_date < datetime.date.today():
        db_plan.status = "Overdue"
    elif db_plan.status == "Overdue" and db_plan.deadline_date >= datetime.date.today():
        db_plan.status = "Planned"
        
    db_plan.updated_at = datetime.datetime.utcnow()
    add_pm_history(db, db_plan, "update", old_data=old_data, note="แก้ไขข้อมูลแผนงาน PM")
    db.commit()
    db.refresh(db_plan)
    return db_plan

@app.post("/api/plans/bulk-delete")
def bulk_delete_plans(payload: BulkDeleteRequest, db: Session = Depends(get_db)):
    if not payload.ids:
        return {"deleted_count": 0}
    plans = db.query(PMPlan).filter(PMPlan.id.in_(payload.ids)).all()
    deleted_count = len(plans)
    for p in plans:
        add_pm_history(db, p, "delete", old_data=_plan_to_dict(p), note="ลบแผนงานแบบหลายรายการ")
        db.delete(p)
    db.commit()
    return {"deleted_count": deleted_count}

@app.delete("/api/plans/{plan_id}")
def delete_plan(plan_id: int, db: Session = Depends(get_db)):
    db_plan = db.query(PMPlan).filter(PMPlan.id == plan_id).first()
    if not db_plan:
        raise HTTPException(status_code=404, detail="Plan not found")
    add_pm_history(db, db_plan, "delete", old_data=_plan_to_dict(db_plan), note="ลบแผนงาน")
    db.delete(db_plan)
    db.commit()
    return {"detail": "Plan deleted successfully"}


# 2. Locations API
@app.get("/api/locations", response_model=List[LocationSchema])
def get_locations(db: Session = Depends(get_db)):
    return db.query(Location).order_by(Location.name.asc()).all()

@app.post("/api/locations", response_model=LocationSchema)
def create_location(loc_in: LocationCreate, db: Session = Depends(get_db)):
    exists = db.query(Location).filter(Location.name == loc_in.name).first()
    if exists:
        raise HTTPException(status_code=400, detail="Location already exists")
    db_loc = Location(name=loc_in.name, province=loc_in.province, district=loc_in.district, service_type=loc_in.service_type, address=loc_in.address, note=loc_in.note)
    db.add(db_loc)
    db.commit()
    db.refresh(db_loc)
    return db_loc

@app.put("/api/locations/{loc_id}", response_model=LocationSchema)
def update_location(loc_id: int, loc_in: LocationCreate, db: Session = Depends(get_db)):
    db_loc = db.query(Location).filter(Location.id == loc_id).first()
    if not db_loc:
        raise HTTPException(status_code=404, detail="Location not found")
    duplicate = db.query(Location).filter(Location.name == loc_in.name, Location.id != loc_id).first()
    if duplicate:
        raise HTTPException(status_code=400, detail="Location already exists")
    db_loc.name = loc_in.name
    db_loc.province = loc_in.province
    db_loc.district = loc_in.district
    db_loc.service_type = loc_in.service_type
    db_loc.address = loc_in.address
    db_loc.note = loc_in.note
    db.commit()
    db.refresh(db_loc)
    return db_loc

@app.delete("/api/locations/{loc_id}")
def delete_location(loc_id: int, db: Session = Depends(get_db)):
    db_loc = db.query(Location).filter(Location.id == loc_id).first()
    if not db_loc:
        raise HTTPException(status_code=404, detail="Location not found")
    db.delete(db_loc)
    db.commit()
    return {"detail": "Location deleted successfully"}



@app.get("/api/locations/export")
def export_locations_csv(db: Session = Depends(get_db)):
    locations = db.query(Location).order_by(Location.name.asc()).all()
    output = io.StringIO()
    output.write("sep=,\n")
    writer = csv.writer(output)
    writer.writerow(["ชื่อศูนย์บริการ/อู่ซ่อม", "ประเภทศูนย์", "จังหวัด", "อำเภอ", "ที่อยู่", "หมายเหตุ"])
    for loc in locations:
        writer.writerow([
            loc.name,
            "ศูนย์ซ่อมภายใน" if loc.service_type == "internal" else "ศูนย์ซ่อมภายนอก",
            loc.province or "",
            loc.district or "",
            loc.address or "",
            loc.note or ""
        ])
    data = b"\xef\xbb\xbf" + output.getvalue().encode("utf-8")
    return StreamingResponse(
        io.BytesIO(data),
        media_type="text/csv",
        headers={"Content-Disposition": "attachment; filename=locations_master_export.csv"}
    )

@app.get("/api/locations/export-xlsx")
def export_locations_xlsx(db: Session = Depends(get_db)):
    locations = db.query(Location).order_by(Location.name.asc()).all()
    wb = Workbook()
    ws = wb.active
    ws.title = "Locations Master"
    headers = ["ชื่อศูนย์บริการ/อู่ซ่อม", "ประเภทศูนย์", "จังหวัด", "อำเภอ", "ที่อยู่", "หมายเหตุ"]
    ws.append(headers)
    for loc in locations:
        ws.append([
            loc.name,
            "ศูนย์ซ่อมภายใน" if loc.service_type == "internal" else "ศูนย์ซ่อมภายนอก",
            loc.province or "",
            loc.district or "",
            loc.address or "",
            loc.note or ""
        ])
    for col in ws.columns:
        max_len = 0
        letter = col[0].column_letter
        for cell in col:
            max_len = max(max_len, len(str(cell.value or "")))
        ws.column_dimensions[letter].width = min(max_len + 3, 55)
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=locations_master_export.xlsx"}
    )

@app.post("/api/locations/import")
async def import_locations_file(file: UploadFile = File(...), db: Session = Depends(get_db)):
    content = await file.read()
    filename = (file.filename or "").lower()

    def rows_from_csv(content_bytes):
        text_data = ""
        for enc in ["utf-8-sig", "cp874", "tis-620", "utf-8", "utf-16"]:
            try:
                text_data = content_bytes.decode(enc)
                break
            except UnicodeDecodeError:
                continue
        if not text_data:
            raise HTTPException(status_code=400, detail="อ่านไฟล์ CSV ไม่ได้ กรุณาใช้ไฟล์ .xlsx หรือ CSV UTF-8/TIS-620")
        lines = text_data.splitlines()
        delimiter = ","
        if lines and lines[0].lower().startswith("sep="):
            delimiter = lines[0][4:5] or ","
            text_data = "\n".join(lines[1:])
        else:
            sample = text_data[:4096]
            counts = {",": sample.count(","), ";": sample.count(";"), "\t": sample.count("\t")}
            delimiter = max(counts, key=counts.get)
            if counts[delimiter] == 0:
                delimiter = ","
        return [[str(cell).strip() for cell in row] for row in csv.reader(io.StringIO(text_data), delimiter=delimiter) if row]

    def rows_from_xlsx(content_bytes):
        try:
            wb = load_workbook(io.BytesIO(content_bytes), data_only=True)
            ws = wb.active
            rows = []
            for row in ws.iter_rows(values_only=True):
                vals = [cell if cell is not None else "" for cell in row]
                if any(str(v).strip() for v in vals):
                    rows.append(vals)
            return rows
        except Exception as e:
            raise HTTPException(status_code=400, detail=f"อ่านไฟล์ Excel ไม่ได้: {str(e)}")

    rows = rows_from_xlsx(content) if filename.endswith(".xlsx") else rows_from_csv(content)
    if not rows:
        raise HTTPException(status_code=400, detail="ไฟล์ที่อัปโหลดไม่มีข้อมูล")
    header = [str(h).strip().lower() for h in rows[0]]
    if len(header) < 2:
        raise HTTPException(status_code=400, detail="ไฟล์มีข้อมูลกองอยู่คอลัมน์เดียว กรุณา Export Excel (.xlsx) แล้วแก้ไขในไฟล์ .xlsx ก่อน Import")

    col_map = {"name": -1, "service_type": -1, "province": -1, "district": -1, "address": -1, "note": -1}
    for idx, h in enumerate(header):
        if "ชื่อ" in h or "ศูนย์" in h or "อู่" in h or "location" in h or "name" in h:
            if col_map["name"] == -1: col_map["name"] = idx
        elif "ประเภท" in h or "service" in h or "type" in h or "ซ่อม" in h:
            if col_map["service_type"] == -1: col_map["service_type"] = idx
        elif "จังหวัด" in h or "province" in h:
            if col_map["province"] == -1: col_map["province"] = idx
        elif "อำเภอ" in h or "district" in h:
            if col_map["district"] == -1: col_map["district"] = idx
        elif "ที่อยู่" in h or "address" in h:
            if col_map["address"] == -1: col_map["address"] = idx
        elif "หมายเหตุ" in h or "note" in h:
            if col_map["note"] == -1: col_map["note"] = idx
    defaults = {"name": 0, "service_type": 1, "province": 2, "district": 3, "address": 4, "note": 5}
    for k, v in defaults.items():
        if col_map[k] == -1:
            col_map[k] = v

    imported_count = 0
    updated_count = 0
    errors = []
    for row_idx, row in enumerate(rows[1:], start=2):
        try:
            if len(row) <= col_map["name"]:
                errors.append(f"แถวที่ {row_idx}: ไม่มีชื่อศูนย์บริการ")
                continue
            name = str(row[col_map["name"]]).strip()
            if not name:
                errors.append(f"แถวที่ {row_idx}: ไม่มีชื่อศูนย์บริการ")
                continue
            raw_type = str(row[col_map["service_type"]]).strip().lower() if col_map["service_type"] < len(row) else "external"
            if raw_type in ["internal", "ซ่อมใน", "ศูนย์ซ่อมภายใน", "ภายใน"] or "ภายใน" in raw_type:
                service_type = "internal"
            else:
                service_type = "external"
            province = str(row[col_map["province"]]).strip() if col_map["province"] < len(row) else ""
            district = str(row[col_map["district"]]).strip() if col_map["district"] < len(row) else ""
            address = str(row[col_map["address"]]).strip() if col_map["address"] < len(row) else ""
            note = str(row[col_map["note"]]).strip() if col_map["note"] < len(row) else ""

            loc = db.query(Location).filter(Location.name == name).first()
            if loc:
                loc.service_type = service_type
                loc.province = province
                loc.district = district
                loc.address = address
                loc.note = note
                updated_count += 1
            else:
                db.add(Location(name=name, service_type=service_type, province=province, district=district, address=address, note=note))
                imported_count += 1
        except Exception as e:
            errors.append(f"แถวที่ {row_idx}: เกิดข้อผิดพลาด {str(e)}")
    add_import_log(db, "locations", filename, "imported", total=max(len(rows)-1, 0), valid=imported_count+updated_count, error=len(errors), imported=imported_count+updated_count, errors=errors)
    db.commit()
    return {"success": True, "imported_count": imported_count, "updated_count": updated_count, "errors": errors}

# 3. Settings API
@app.get("/api/settings", response_model=SettingsSchema)
def get_settings(db: Session = Depends(get_db)):
    token = db.query(Setting).filter(Setting.key == "line_channel_access_token").first()
    target = db.query(Setting).filter(Setting.key == "line_target_id").first()
    days = db.query(Setting).filter(Setting.key == "alert_days_before").first()
    time_val = db.query(Setting).filter(Setting.key == "scheduler_time").first()
    enabled = db.query(Setting).filter(Setting.key == "scheduler_enabled").first()
    template = db.query(Setting).filter(Setting.key == "notification_template").first()
    secret = db.query(Setting).filter(Setting.key == "line_channel_secret").first()
    webhook_enabled = db.query(Setting).filter(Setting.key == "line_webhook_enabled").first()
    debug_mode = db.query(Setting).filter(Setting.key == "debug_mode").first()
    
    return SettingsSchema(
        line_channel_access_token=token.value if token else "",
        line_target_id=target.value if target else "",
        line_channel_secret=secret.value if secret else "",
        line_webhook_enabled=(webhook_enabled.value.lower() == "true") if webhook_enabled and webhook_enabled.value else True,
        alert_days_before=int(days.value) if days and days.value else 3,
        scheduler_time=time_val.value if time_val else "08:00",
        scheduler_enabled=(enabled.value.lower() == "true") if enabled else False,
        notification_template=template.value if template else "",
        debug_mode=(debug_mode.value.lower() == "true") if debug_mode and debug_mode.value else False,
        report_morning_time=_get_setting_value(db, "report_morning_time", "07:30"),
        report_noon_time=_get_setting_value(db, "report_noon_time", "12:00"),
        report_evening_time=_get_setting_value(db, "report_evening_time", "17:00"),
        report_morning_enabled=_get_setting_value(db, "report_morning_enabled", "true").lower()=="true",
        report_noon_enabled=_get_setting_value(db, "report_noon_enabled", "true").lower()=="true",
        report_evening_enabled=_get_setting_value(db, "report_evening_enabled", "true").lower()=="true"
    )

@app.post("/api/settings")
def update_settings(settings_in: SettingsSchema, db: Session = Depends(get_db)):
    db_token = db.query(Setting).filter(Setting.key == "line_channel_access_token").first()
    if db_token: db_token.value = settings_in.line_channel_access_token
    
    db_target = db.query(Setting).filter(Setting.key == "line_target_id").first()
    if db_target: db_target.value = settings_in.line_target_id

    db_secret = db.query(Setting).filter(Setting.key == "line_channel_secret").first()
    if db_secret:
        db_secret.value = settings_in.line_channel_secret or ""
    else:
        db.add(Setting(key="line_channel_secret", value=settings_in.line_channel_secret or ""))

    db_webhook_enabled = db.query(Setting).filter(Setting.key == "line_webhook_enabled").first()
    if db_webhook_enabled:
        db_webhook_enabled.value = str(settings_in.line_webhook_enabled).lower()
    else:
        db.add(Setting(key="line_webhook_enabled", value=str(settings_in.line_webhook_enabled).lower()))
    
    db_days = db.query(Setting).filter(Setting.key == "alert_days_before").first()
    if db_days: db_days.value = str(settings_in.alert_days_before)
    
    db_time = db.query(Setting).filter(Setting.key == "scheduler_time").first()
    if db_time: db_time.value = settings_in.scheduler_time
    
    db_enabled = db.query(Setting).filter(Setting.key == "scheduler_enabled").first()
    if db_enabled: db_enabled.value = str(settings_in.scheduler_enabled).lower()

    db_template = db.query(Setting).filter(Setting.key == "notification_template").first()
    if db_template:
        db_template.value = settings_in.notification_template or ""
    else:
        db.add(Setting(key="notification_template", value=settings_in.notification_template or ""))
    
    db_debug = db.query(Setting).filter(Setting.key == "debug_mode").first()
    if db_debug:
        db_debug.value = str(settings_in.debug_mode).lower()
    else:
        db.add(Setting(key="debug_mode", value=str(settings_in.debug_mode).lower()))
    for key, value in {
        "report_morning_time": settings_in.report_morning_time or "07:30",
        "report_noon_time": settings_in.report_noon_time or "12:00",
        "report_evening_time": settings_in.report_evening_time or "17:00",
        "report_morning_enabled": str(bool(settings_in.report_morning_enabled)).lower(),
        "report_noon_enabled": str(bool(settings_in.report_noon_enabled)).lower(),
        "report_evening_enabled": str(bool(settings_in.report_evening_enabled)).lower(),
    }.items():
        row=db.query(Setting).filter(Setting.key==key).first()
        if row: row.value=str(value)
        else: db.add(Setting(key=key,value=str(value)))
    db.commit()
    update_scheduler_job(db)
    update_v161_report_scheduler(db)
    # Auto Test after Save: returns validation only, not push, to avoid accidental spam.
    validation = validate_line_config(settings_in.line_channel_access_token, settings_in.line_target_id, settings_in.scheduler_time)
    write_app_log("system.log", json.dumps({"event":"settings_saved", "validation":validation}, ensure_ascii=False))
    return {"detail": "Settings updated successfully", "validation": validation}


def validate_line_config(token: str, target_id: str, scheduler_time: str = "08:00"):
    checks = []
    def add(name, ok, message):
        checks.append({"name": name, "ok": bool(ok), "message": message})
    add("token", bool(token and len(token.strip()) > 40), "Token ดูปกติ" if token and len(token.strip()) > 40 else "Token ว่างหรือสั้นเกินไป")
    add("target", bool(target_id and target_id[:1] in ["C","U","R"]), "Target ID format ถูกต้อง" if target_id and target_id[:1] in ["C","U","R"] else "Target ID ต้องขึ้นต้นด้วย C/U/R")
    try:
        hh, mm = scheduler_time.split(":")
        ok = 0 <= int(hh) <= 23 and 0 <= int(mm) <= 59
    except Exception:
        ok = False
    add("scheduler_time", ok, "รูปแบบเวลาถูกต้อง" if ok else "รูปแบบเวลาต้องเป็น HH:MM")
    return {"ok": all(c["ok"] for c in checks), "checks": checks}

class TestLinePayload(BaseModel):
    token: str
    target_id: str
    message: str

@app.post("/api/settings/test-line")
def test_line_api(payload: TestLinePayload, db: Session = Depends(get_db)):
    result = send_line_message_debug(payload.token, payload.target_id, payload.message)
    log = NotificationLog(channel="LINE", target_id=payload.target_id, message=payload.message, status="success" if result.get("success") else "failed", response=json.dumps(result, ensure_ascii=False))
    db.add(log)
    setting = db.query(Setting).filter(Setting.key == "last_line_debug").first()
    if setting:
        setting.value = json.dumps(result, ensure_ascii=False)
    else:
        db.add(Setting(key="last_line_debug", value=json.dumps(result, ensure_ascii=False)))
    db.commit()
    write_app_log("line.log", json.dumps(result, ensure_ascii=False))
    # Return 200 even when LINE rejects, so UI can show structured analyzer instead of generic 500.
    return result

@app.post("/api/settings/trigger-daily")
def trigger_daily_api(db: Session = Depends(get_db)):
    success = trigger_daily_notification(db)
    if success:
        return {"status": "success", "message": "Daily report sent to Line successfully."}
    else:
        raise HTTPException(status_code=500, detail="Failed to send report. Ensure settings are saved and Line parameters are valid.")



# 3.1 LINE Webhook / Integration Center

def _get_setting_value(db: Session, key: str, default: str = "") -> str:
    row = db.query(Setting).filter(Setting.key == key).first()
    return row.value if row and row.value is not None else default

def _verify_line_signature(channel_secret: str, body: bytes, signature: Optional[str]) -> bool:
    if not channel_secret:
        return True  # Signature check is optional while developing locally.
    if not signature:
        return False
    digest = hmac.new(channel_secret.encode("utf-8"), body, hashlib.sha256).digest()
    expected = base64.b64encode(digest).decode("utf-8")
    return hmac.compare_digest(expected, signature)

def _line_reply(token: str, reply_token: str, text: str) -> bool:
    if not token or not reply_token:
        return False
    try:
        response = requests.post(
            "https://api.line.me/v2/bot/message/reply",
            headers={"Content-Type": "application/json", "Authorization": f"Bearer {token}"},
            json={"replyToken": reply_token, "messages": [{"type": "text", "text": text[:4900]}]},
            timeout=10,
        )
        return response.status_code == 200
    except Exception as e:
        print("LINE reply error:", e)
        return False

@app.post("/line/webhook")
async def line_webhook(request: Request, x_line_signature: Optional[str] = Header(None), db: Session = Depends(get_db)):
    """Webhook endpoint for LINE OA. Use with ngrok: https://xxxxx.ngrok-free.app/line/webhook"""
    enabled = _get_setting_value(db, "line_webhook_enabled", "true").lower() == "true"
    if not enabled:
        return {"status": "disabled"}

    body = await request.body()
    channel_secret = _get_setting_value(db, "line_channel_secret", "")
    if not _verify_line_signature(channel_secret, body, x_line_signature):
        db.add(LineWebhookEvent(event_type="signature_error", raw_payload=body.decode("utf-8", errors="ignore"), status="rejected"))
        db.commit()
        raise HTTPException(status_code=401, detail="Invalid LINE signature")

    try:
        payload = json.loads(body.decode("utf-8")) if body else {}
    except Exception:
        db.add(LineWebhookEvent(event_type="json_error", raw_payload=body.decode("utf-8", errors="ignore"), status="failed"))
        db.commit()
        raise HTTPException(status_code=400, detail="Invalid JSON")

    token = _get_setting_value(db, "line_channel_access_token", "")
    events = payload.get("events", [])
    discovered = []

    for event in events:
        source = event.get("source", {}) or {}
        source_type = source.get("type") or "unknown"
        source_id = source.get("groupId") or source.get("userId") or source.get("roomId") or ""
        event_type = event.get("type") or "unknown"
        raw_text = json.dumps(event, ensure_ascii=False)

        db.add(LineWebhookEvent(event_type=event_type, source_type=source_type, source_id=source_id, raw_payload=raw_text, status="received"))
        write_app_log("webhook.log", raw_text)

        if source_id:
            target = db.query(LineTarget).filter(LineTarget.source_id == source_id).first()
            if not target:
                label_map = {"group": "LINE Group", "user": "LINE User", "room": "LINE Room"}
                target = LineTarget(source_type=source_type, source_id=source_id, display_name=f"{label_map.get(source_type, 'LINE Target')} {source_id[-6:]}")
                db.add(target)
            target.source_type = source_type
            target.is_active = True
            target.last_event_at = datetime.datetime.utcnow()
            discovered.append({"type": source_type, "id": source_id})

        # Friendly command reply. This also proves webhook works and shows the Target ID.
        reply_token = event.get("replyToken")
        message = event.get("message", {}) or {}
        text_message = (message.get("text") or "").strip().lower()
        if reply_token and text_message in ["id", "groupid", "group id", "target", "ทดสอบ", "test"]:
            if source_id:
                _line_reply(token, reply_token, f"✅ PM Tracking รับ Webhook แล้ว\nType: {source_type}\nTarget ID: {source_id}\n\nนำ Target ID นี้ไปใช้ส่ง Push Message ได้ครับ")
            else:
                _line_reply(token, reply_token, "✅ PM Tracking รับ Webhook แล้ว แต่ไม่พบ source id")

    db.commit()
    return {"status": "ok", "events": len(events), "discovered": discovered}

@app.get("/api/line/targets", response_model=List[LineTargetSchema])
def get_line_targets(db: Session = Depends(get_db)):
    return db.query(LineTarget).order_by(LineTarget.last_event_at.desc()).all()

@app.post("/api/line/targets/use")
def use_line_target(payload: UseLineTargetPayload, db: Session = Depends(get_db)):
    target = db.query(LineTarget).filter(LineTarget.source_id == payload.target_id).first()
    if not target:
        raise HTTPException(status_code=404, detail="Line target not found")
    setting = db.query(Setting).filter(Setting.key == "line_target_id").first()
    if setting:
        setting.value = payload.target_id
    else:
        db.add(Setting(key="line_target_id", value=payload.target_id))
    db.commit()
    update_scheduler_job(db)
    return {"status": "success", "target_id": payload.target_id}

@app.get("/api/line/webhook-info")
def line_webhook_info(request: Request):
    return {"webhook_path": "/line/webhook", "current_base_url": str(request.base_url).rstrip("/"), "webhook_url": str(request.base_url).rstrip("/") + "/line/webhook"}

@app.get("/api/line/debug/status")
def line_debug_status(request: Request, db: Session = Depends(get_db)):
    token = _get_setting_value(db, "line_channel_access_token", "")
    target_id = _get_setting_value(db, "line_target_id", "")
    sched_time = _get_setting_value(db, "scheduler_time", "08:00")
    last_debug = _get_setting_value(db, "last_line_debug", "")
    targets_count = db.query(LineTarget).count()
    webhook_count = db.query(LineWebhookEvent).count()
    notif_count = db.query(NotificationLog).count()
    last_webhook = db.query(LineWebhookEvent).order_by(LineWebhookEvent.received_at.desc()).first()
    last_notif = db.query(NotificationLog).order_by(NotificationLog.sent_at.desc()).first()
    validation = validate_line_config(token, target_id, sched_time)
    return {
        "version": "v1.4.1 LINE Diagnostic Build",
        "connection": validation,
        "token_masked": token[:8] + "..." + token[-8:] if token and len(token) > 20 else "",
        "target_id": target_id,
        "target_type": "group" if target_id.startswith("C") else ("user" if target_id.startswith("U") else ("room" if target_id.startswith("R") else "unknown")),
        "webhook_url": str(request.base_url).rstrip("/") + "/line/webhook",
        "targets_count": targets_count,
        "webhook_events_count": webhook_count,
        "notification_logs_count": notif_count,
        "last_webhook": {"event_type": last_webhook.event_type, "source_type": last_webhook.source_type, "source_id": last_webhook.source_id, "received_at": last_webhook.received_at.isoformat(), "status": last_webhook.status} if last_webhook else None,
        "last_notification": {"status": last_notif.status, "target_id": last_notif.target_id, "sent_at": last_notif.sent_at.isoformat()} if last_notif else None,
        "last_debug": json.loads(last_debug) if last_debug else None,
        "scheduler_enabled": _get_setting_value(db, "scheduler_enabled", "false"),
        "debug_mode": _get_setting_value(db, "debug_mode", "false"),
    }

@app.get("/api/line/inspector")
def line_inspector(request: Request, db: Session = Depends(get_db)):
    """LINE Inspector: one-stop diagnostic summary for v1.4.1."""
    token = _get_setting_value(db, "line_channel_access_token", "")
    target_id = _get_setting_value(db, "line_target_id", "")
    secret = _get_setting_value(db, "line_channel_secret", "")
    webhook_enabled = _get_setting_value(db, "line_webhook_enabled", "true").lower() == "true"
    last_webhook = db.query(LineWebhookEvent).order_by(LineWebhookEvent.received_at.desc()).first()
    last_debug_raw = _get_setting_value(db, "last_line_debug", "")
    try:
        last_debug = json.loads(last_debug_raw) if last_debug_raw else None
    except Exception:
        last_debug = {"raw": last_debug_raw}
    items = [
        {"name": "Channel Access Token", "ok": bool(token and len(token) > 40), "detail": (token[:8] + "..." + token[-8:]) if token and len(token) > 20 else "Missing"},
        {"name": "Channel Secret", "ok": bool(secret), "detail": "Configured" if secret else "Optional / not set"},
        {"name": "Webhook Switch", "ok": webhook_enabled, "detail": "ON" if webhook_enabled else "OFF"},
        {"name": "Webhook URL", "ok": True, "detail": str(request.base_url).rstrip("/") + "/line/webhook"},
        {"name": "Target Selected", "ok": bool(target_id and target_id[0:1] in ["C", "U", "R"]), "detail": target_id or "Missing"},
        {"name": "Targets Discovered", "ok": db.query(LineTarget).count() > 0, "detail": f"{db.query(LineTarget).count()} targets"},
        {"name": "Last Webhook", "ok": bool(last_webhook), "detail": last_webhook.received_at.isoformat() if last_webhook else "No webhook events"},
        {"name": "Last Push", "ok": bool(last_debug and last_debug.get("success")), "detail": ("Success" if last_debug and last_debug.get("success") else (last_debug.get("analysis",{}).get("root_cause") if last_debug else "Not tested"))},
    ]
    return {"ok": all(i["ok"] for i in items if i["name"] not in ["Channel Secret"]), "items": items, "last_debug": last_debug}

@app.get("/api/line/request-history")
def line_request_history(limit: int = 50, db: Session = Depends(get_db)):
    rows = db.query(NotificationLog).filter(NotificationLog.channel == "LINE").order_by(NotificationLog.sent_at.desc()).limit(limit).all()
    out = []
    for r in rows:
        try:
            debug = json.loads(r.response or "{}")
        except Exception:
            debug = {"response_text": r.response}
        out.append({
            "id": r.id,
            "sent_at": r.sent_at.isoformat(),
            "target_id": r.target_id,
            "status": r.status,
            "status_code": debug.get("status_code"),
            "root_cause": (debug.get("analysis") or {}).get("root_cause") or (debug.get("analysis") or {}).get("reason"),
            "line_message": (debug.get("analysis") or {}).get("line_message") or debug.get("response_text"),
            "elapsed_ms": debug.get("elapsed_ms"),
            "debug": debug,
        })
    return out

@app.get("/api/line/webhook-events")
def get_webhook_events(db: Session = Depends(get_db)):
    rows = db.query(LineWebhookEvent).order_by(LineWebhookEvent.received_at.desc()).limit(50).all()
    return [{"id": r.id, "event_type": r.event_type, "source_type": r.source_type, "source_id": r.source_id, "status": r.status, "received_at": r.received_at.isoformat(), "raw_payload": r.raw_payload} for r in rows]

@app.get("/api/line/error-analyzer")
def get_last_line_error(db: Session = Depends(get_db)):
    row = db.query(NotificationLog).order_by(NotificationLog.sent_at.desc()).first()
    if not row:
        return {"message": "ยังไม่มี Notification Log"}
    try:
        data = json.loads(row.response or "{}")
    except Exception:
        data = {"response_text": row.response}
    return {"status": row.status, "sent_at": row.sent_at.isoformat(), "target_id": row.target_id, "message": row.message, "debug": data}

@app.get("/api/system/health")
def system_health(db: Session = Depends(get_db)):
    token = _get_setting_value(db, "line_channel_access_token", "")
    target_id = _get_setting_value(db, "line_target_id", "")
    checks = []
    def add(name, ok, detail, category="system"):
        checks.append({"name": name, "ok": bool(ok), "detail": detail, "category": category})
    add("FastAPI", True, "Server running", "server")
    try:
        db.query(Setting).count(); add("SQLite", True, "Database connection OK", "database")
    except Exception as e:
        add("SQLite", False, str(e), "database")
    add("Vehicle_Master", db.query(VehicleMaster).count() > 0, f"{db.query(VehicleMaster).count()} vehicles", "database")
    add("PM_Plan", db.query(PMPlan).count() >= 0, f"{db.query(PMPlan).count()} plans", "database")
    add("LINE Token", bool(token and len(token) > 40), "Configured" if token else "Missing", "line")
    add("LINE Target", bool(target_id and target_id[:1] in ["C","U","R"]), target_id or "Missing", "line")
    add("Webhook Events", db.query(LineWebhookEvent).count() > 0, f"{db.query(LineWebhookEvent).count()} events", "line")
    add("LINE Targets", db.query(LineTarget).count() > 0, f"{db.query(LineTarget).count()} targets", "line")
    add("Log Folder", LOG_DIR.exists(), str(LOG_DIR), "logs")
    for log in ["line.log", "webhook.log", "scheduler.log", "system.log", "api.log"]:
        p = LOG_DIR / log
        if not p.exists():
            try: p.touch()
            except Exception: pass
        add(log, p.exists(), f"{p.stat().st_size if p.exists() else 0} bytes", "logs")
    return {"ok": all(c["ok"] for c in checks), "checks": checks, "checked_at": datetime.datetime.now().isoformat(timespec='seconds')}

@app.get("/api/system/logs/{log_name}")
def read_system_log(log_name: str):
    allowed = {"line", "scheduler", "webhook", "system", "api"}
    if log_name not in allowed:
        raise HTTPException(status_code=400, detail="Invalid log name")
    p = LOG_DIR / f"{log_name}.log"
    if not p.exists():
        return {"log_name": log_name, "lines": []}
    lines = p.read_text(encoding="utf-8", errors="ignore").splitlines()[-200:]
    return {"log_name": log_name, "lines": lines}

class ApiTesterPayload(BaseModel):
    method: str = "GET"
    path: str
    body: Optional[dict] = None

@app.post("/api/dev/api-test")
def api_tester(payload: ApiTesterPayload, request: Request):
    """Safe API Tester: returns the request that would be executed for local diagnostics."""
    path = payload.path if payload.path.startswith("/") else "/" + payload.path
    result = {
        "method": payload.method.upper(),
        "path": path,
        "base_url": str(request.base_url).rstrip("/"),
        "body": payload.body or {},
        "note": "API Tester v1.4.1: ใช้สำหรับเตรียม/ตรวจ payload ภายในระบบ ไม่ยิง external API โดยตรง"
    }
    write_app_log("api.log", json.dumps({"event":"api_tester", **result}, ensure_ascii=False))
    return result

@app.get("/api/system/snapshot")
def system_snapshot(db: Session = Depends(get_db)):
    """Create a ZIP snapshot for debug without exposing full LINE token."""
    snap_id = datetime.datetime.now().strftime("%Y%m%d_%H%M%S") + "_" + uuid.uuid4().hex[:6]
    tmp_root = Path(__file__).resolve().parent / "_snapshots"
    tmp_root.mkdir(exist_ok=True)
    snap_dir = tmp_root / f"debug_snapshot_{snap_id}"
    snap_dir.mkdir(exist_ok=True)
    token = _get_setting_value(db, "line_channel_access_token", "")
    target_id = _get_setting_value(db, "line_target_id", "")
    info = {
        "version": "PM Tracking System v1.4.1 LINE Diagnostic Build",
        "created_at": datetime.datetime.now().isoformat(timespec='seconds'),
        "python": platform.python_version(),
        "platform": platform.platform(),
        "db_counts": {
            "vehicles": db.query(VehicleMaster).count(),
            "locations": db.query(Location).count(),
            "pm_plans": db.query(PMPlan).count(),
            "line_targets": db.query(LineTarget).count(),
            "webhook_events": db.query(LineWebhookEvent).count(),
            "notification_logs": db.query(NotificationLog).count(),
        },
        "settings_safe": {
            "line_channel_access_token_masked": token[:8] + "..." + token[-8:] if token and len(token) > 20 else "",
            "line_target_id": target_id,
            "scheduler_time": _get_setting_value(db, "scheduler_time", ""),
            "scheduler_enabled": _get_setting_value(db, "scheduler_enabled", ""),
            "debug_mode": _get_setting_value(db, "debug_mode", ""),
        }
    }
    (snap_dir / "system_info.json").write_text(json.dumps(info, ensure_ascii=False, indent=2), encoding="utf-8")
    for name in ["line", "webhook", "scheduler", "system", "api"]:
        p = LOG_DIR / f"{name}.log"
        (snap_dir / f"{name}.log").write_text(p.read_text(encoding="utf-8", errors="ignore") if p.exists() else "", encoding="utf-8")
    last_debug = _get_setting_value(db, "last_line_debug", "")
    (snap_dir / "last_line_debug.json").write_text(last_debug or "{}", encoding="utf-8")
    zip_path = tmp_root / f"debug_snapshot_{snap_id}.zip"
    with zipfile.ZipFile(zip_path, "w", zipfile.ZIP_DEFLATED) as z:
        for f in snap_dir.rglob("*"):
            z.write(f, f.relative_to(snap_dir))
    shutil.rmtree(snap_dir, ignore_errors=True)
    return FileResponse(str(zip_path), filename=zip_path.name, media_type="application/zip")

@app.post("/api/line/simulator")
def line_message_simulator(db: Session = Depends(get_db)):
    text = get_alert_summary_text(db)
    target_id = _get_setting_value(db, "line_target_id", "")
    payload = {"to": target_id, "messages": [{"type": "text", "text": text}]}
    return {"target_id": target_id, "message": text, "payload": payload}


# 4. Summary & Dashboard API
@app.get("/api/summary")
def get_dashboard_summary(db: Session = Depends(get_db)):
    today = datetime.date.today()
    tomorrow = today + datetime.timedelta(days=1)
    
    days_setting = db.query(Setting).filter(Setting.key == "alert_days_before").first()
    alert_days = int(days_setting.value) if days_setting and days_setting.value else 3
    upcoming_limit = today + datetime.timedelta(days=alert_days)
    
    total_vehicles = db.query(PMPlan.vehicle_no).distinct().count()
    total_plans = db.query(PMPlan).filter(PMPlan.status != "Cancelled").count()
    
    planned = db.query(PMPlan).filter(PMPlan.status == "Planned").count()
    in_progress = db.query(PMPlan).filter(PMPlan.status == "In Progress").count()
    completed = db.query(PMPlan).filter(PMPlan.status == "Completed").count()
    cancelled = db.query(PMPlan).filter(PMPlan.status == "Cancelled").count()
    
    overdue = db.query(PMPlan).filter(
        PMPlan.deadline_date < today,
        PMPlan.status != "Completed",
        PMPlan.status != "Cancelled"
    ).count()
    
    near_due = db.query(PMPlan).filter(
        PMPlan.planned_date >= today,
        PMPlan.planned_date <= upcoming_limit,
        PMPlan.status != "Completed",
        PMPlan.status != "Cancelled"
    ).count()
    
    tomorrow_plans = db.query(PMPlan).filter(
        PMPlan.planned_date == tomorrow,
        PMPlan.status != "Cancelled"
    ).count()

    location_stats = {}
    province_stats = {}
    service_type_stats = {"internal": 0, "external": 0}
    locations = db.query(Location).all()
    for loc in locations:
        loc_name = loc.name
        count = db.query(PMPlan).filter(PMPlan.location == loc_name, PMPlan.status != "Cancelled").count()
        location_stats[loc_name] = count
        if loc.province:
            province_stats[loc.province] = province_stats.get(loc.province, 0) + count
        service_type_stats[loc.service_type or "external"] = service_type_stats.get(loc.service_type or "external", 0) + count

    transport_stats = {}
    for row in db.query(PMPlan.transport_type).filter(PMPlan.status != "Cancelled").all():
        key = row[0] or "ไม่ระบุ"
        transport_stats[key] = transport_stats.get(key, 0) + 1

    job_stats = {}
    for row in db.query(PMPlan.job_title).filter(PMPlan.status != "Cancelled").all():
        key = row[0] or "ไม่ระบุชื่องาน"
        job_stats[key] = job_stats.get(key, 0) + 1

    due_today = db.query(PMPlan).filter(PMPlan.planned_date == today, PMPlan.status != "Cancelled").count()
    month_start = today.replace(day=1)
    next_month = (month_start.replace(year=month_start.year + 1, month=1) if month_start.month == 12 else month_start.replace(month=month_start.month + 1))
    this_month = db.query(PMPlan).filter(PMPlan.planned_date >= month_start, PMPlan.planned_date < next_month, PMPlan.status != "Cancelled").count()
    completion_rate = round((completed / total_plans) * 100, 1) if total_plans else 0

    upcoming_list = db.query(PMPlan).filter(
        PMPlan.planned_date >= today,
        PMPlan.status != "Completed",
        PMPlan.status != "Cancelled"
    ).order_by(PMPlan.planned_date.asc()).limit(5).all()
    
    upcoming_parsed = []
    for plan in upcoming_list:
        upcoming_parsed.append({
            "id": plan.id,
            "vehicle_no": plan.vehicle_no,
            "job_title": plan.job_title,
            "planned_date": plan.planned_date.strftime("%Y-%m-%d"),
            "deadline_date": plan.deadline_date.strftime("%Y-%m-%d"),
            "location": plan.location,
            "transport_type": plan.transport_type,
            "vehicle_model": plan.vehicle_model,
            "fleet": plan.fleet,
            "status": plan.status
        })

    return {
        "total_vehicles": total_vehicles,
        "total_plans": total_plans,
        "planned": planned,
        "in_progress": in_progress,
        "completed": completed,
        "cancelled": cancelled,
        "overdue": overdue,
        "near_due": near_due,
        "tomorrow_plans": tomorrow_plans,
        "due_today": due_today,
        "this_month": this_month,
        "completion_rate": completion_rate,
        "location_stats": location_stats,
        "province_stats": province_stats,
        "service_type_stats": service_type_stats,
        "transport_stats": transport_stats,
        "job_stats": job_stats,
        "upcoming_list": upcoming_parsed
    }


# 5. Car Data CSV API (Load Data Car.csv)
@app.get("/api/cars")
def get_cars_from_csv():
    """
    Reads the 'Data Car.csv' file in the workspace directory (TIS-620 / CP874 encoding)
    and returns a structured list of cars for dropdown suggest.
    """
    csv_path = os.path.join(os.path.dirname(__file__), "Data Car.csv")
    if not os.path.exists(csv_path):
        # Fallback to current directory or parent directory
        csv_path = "Data Car.csv"
        
    if not os.path.exists(csv_path):
        return []
        
    cars = []
    try:
        with open(csv_path, mode="r", encoding="cp874", errors="ignore") as f:
            reader = csv.reader(f)
            # Skip header row
            try:
                header = next(reader)
            except StopIteration:
                return []
                
            for row in reader:
                if len(row) >= 5:
                    # Column A: รหัสรถ, Column B: ประเภทขนส่ง, Column C: รุ่นรถ, Column D: ประจำคลังโยกฟลีต, Column E: เบอร์รถ
                    car_code = row[0].strip()
                    transport_type = row[1].strip()
                    car_model = row[2].strip()
                    fleet = row[3].strip()
                    car_no = row[4].strip()
                    
                    vehicle_type = row[5].strip() if len(row) >= 6 else ""
                    if car_no:
                        # Construct a search label: [เบอร์รถ] | รหัสรถ | รุ่นรถ | ประเภทขนส่ง | ประจำคลังโยกฟลีต
                        label = f"{car_no} | {car_code} | {car_model} | {transport_type} | {fleet}"
                        cars.append({
                            "car_no": car_no,
                            "code": car_code,
                            "model": car_model,
                            "type": transport_type,
                            "fleet": fleet,
                            "vehicle_type": vehicle_type,
                            "label": label
                        })
    except Exception as e:
        print("Error reading Data Car.csv:", e)
        return []
        
    # Sort by vehicle number
    return sorted(cars, key=lambda k: k["car_no"])


# 6. PM Plans Import / Export APIs
@app.get("/api/plans/export")
def export_plans_csv(db: Session = Depends(get_db)):
    """
    Exports all PM plans from database as a CSV file.
    Prefixed with UTF-8 BOM so Excel opens it with correct Thai characters.
    """
    plans = db.query(PMPlan).order_by(PMPlan.planned_date.desc()).all()
    
    output = io.StringIO()
    writer = csv.writer(output)
    
    # Excel hint: make Excel split columns correctly even on Thai/Windows locale
    output.write("sep=,\n")

    # Write Header row
    writer.writerow([
        "ทะเบียนรถ/เบอร์รถ", "ชื่องาน PM", "รายละเอียด", "วันที่นัดเข้า PM", 
        "วันที่เข้าทำจริง", "วัน Deadline", "สถานที่นัดเข้า", "สถานะ"
    ])
    
    # Write Data rows
    for p in plans:
        planned_str = p.planned_date.strftime("%Y-%m-%d") if p.planned_date else ""
        actual_str = p.actual_date.strftime("%Y-%m-%d") if p.actual_date else ""
        deadline_str = p.deadline_date.strftime("%Y-%m-%d") if p.deadline_date else ""
        
        writer.writerow([
            p.vehicle_no, p.job_title, p.description or "", planned_str,
            actual_str, deadline_str, p.location, p.status
        ])
        
    # Get string and prefix with UTF-8 BOM
    csv_data = output.getvalue()
    bom_utf8_csv = b'\xef\xbb\xbf' + csv_data.encode("utf-8")
    
    return StreamingResponse(
        io.BytesIO(bom_utf8_csv),
        media_type="text/csv",
        headers={"Content-Disposition": "attachment; filename=pm_plans_export.csv"}
    )


@app.get("/api/plans/export-xlsx")
def export_plans_xlsx(db: Session = Depends(get_db)):
    """
    Exports all PM plans as a real Excel .xlsx file.
    This avoids the common Thai Windows/Excel CSV issue where all data appears in one column.
    """
    plans = db.query(PMPlan).order_by(PMPlan.planned_date.desc()).all()

    wb = Workbook()
    ws = wb.active
    ws.title = "PM Plans"

    headers = [
        "ทะเบียนรถ/เบอร์รถ", "ชื่องาน PM", "รายละเอียด", "วันที่นัดเข้า PM",
        "วันที่เข้าทำจริง", "วัน Deadline", "สถานที่นัดเข้า", "สถานะ"
    ]
    ws.append(headers)

    for p in plans:
        ws.append([
            p.vehicle_no,
            p.job_title,
            p.description or "",
            p.planned_date.strftime("%Y-%m-%d") if p.planned_date else "",
            p.actual_date.strftime("%Y-%m-%d") if p.actual_date else "",
            p.deadline_date.strftime("%Y-%m-%d") if p.deadline_date else "",
            p.location,
            p.status
        ])

    for col in ws.columns:
        max_len = 0
        letter = col[0].column_letter
        for cell in col:
            max_len = max(max_len, len(str(cell.value or "")))
        ws.column_dimensions[letter].width = min(max_len + 3, 45)

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=pm_plans_export.xlsx"}
    )

@app.post("/api/plans/import")
async def import_plans_file(file: UploadFile = File(...), db: Session = Depends(get_db)):
    """
    Imports PM plans from CSV or real Excel .xlsx.
    Recommended file type is .xlsx because Excel often breaks CSV into one long column on Thai Windows locale.
    """
    content = await file.read()
    filename = (file.filename or "").lower()

    def parse_date(date_value):
        if date_value is None:
            return None

        if isinstance(date_value, datetime.datetime):
            return date_value.date()
        if isinstance(date_value, datetime.date):
            return date_value

        date_str = str(date_value).strip()
        if not date_str:
            return None

        parsed_date = None
        for fmt in ["%Y-%m-%d", "%d/%m/%Y", "%Y/%m/%d", "%d-%m-%Y", "%d/%m/%y", "%d-%m-%y"]:
            try:
                parsed_date = datetime.datetime.strptime(date_str, fmt).date()
                break
            except ValueError:
                continue

        if parsed_date and parsed_date.year > 2400:
            try:
                parsed_date = parsed_date.replace(year=parsed_date.year - 543)
            except ValueError:
                parsed_date = parsed_date.replace(year=parsed_date.year - 543, day=28)

        return parsed_date

    def normalize_rows_from_csv(content_bytes):
        text_data = ""
        for enc in ["utf-8-sig", "cp874", "tis-620", "utf-8", "utf-16"]:
            try:
                text_data = content_bytes.decode(enc)
                break
            except UnicodeDecodeError:
                continue

        if not text_data:
            raise HTTPException(status_code=400, detail="อ่านไฟล์ CSV ไม่ได้ กรุณาใช้ไฟล์ .xlsx หรือ CSV UTF-8/TIS-620")

        # Support Excel delimiter hint e.g. first line: sep=,
        lines = text_data.splitlines()
        delimiter = ","
        if lines and lines[0].lower().startswith("sep="):
            delimiter = lines[0][4:5] or ","
            text_data = "\n".join(lines[1:])
        else:
            sample = text_data[:4096]
            counts = {",": sample.count(","), ";": sample.count(";"), "\t": sample.count("\t")}
            delimiter = max(counts, key=counts.get)
            if counts[delimiter] == 0:
                delimiter = ","

        reader = csv.reader(io.StringIO(text_data), delimiter=delimiter)
        return [[str(cell).strip() for cell in row] for row in reader if row]

    def normalize_rows_from_xlsx(content_bytes):
        try:
            wb = load_workbook(io.BytesIO(content_bytes), data_only=True)
            ws = wb.active
            rows = []
            for row in ws.iter_rows(values_only=True):
                values = [cell if cell is not None else "" for cell in row]
                if any(str(v).strip() for v in values):
                    rows.append(values)
            return rows
        except Exception as e:
            raise HTTPException(status_code=400, detail=f"อ่านไฟล์ Excel ไม่ได้: {str(e)}")

    if filename.endswith(".xlsx"):
        rows = normalize_rows_from_xlsx(content)
    else:
        rows = normalize_rows_from_csv(content)

    if not rows:
        raise HTTPException(status_code=400, detail="ไฟล์ที่อัปโหลดไม่มีข้อมูล")

    header = [str(h).strip().lower() for h in rows[0]]

    # If CSV was saved/opened until every row became one long column, import cannot know the original columns.
    if len(header) < 2:
        raise HTTPException(
            status_code=400,
            detail="ไฟล์มีข้อมูลกองอยู่คอลัมน์เดียว ทำให้นำเข้าไม่ได้ กรุณากด Export Excel (.xlsx) แล้วแก้ไขในไฟล์ .xlsx ก่อน Import"
        )

    col_map = {
        "vehicle_no": -1,
        "job_title": -1,
        "description": -1,
        "planned_date": -1,
        "actual_date": -1,
        "deadline_date": -1,
        "location": -1,
        "status": -1
    }

    for idx, h in enumerate(header):
        if "ทะเบียน" in h or "vehicle" in h or "เบอร์รถ" in h:
            if col_map["vehicle_no"] == -1: col_map["vehicle_no"] = idx
        elif "ชื่องาน" in h or "job" in h or "งาน pm" in h:
            if col_map["job_title"] == -1: col_map["job_title"] = idx
        elif "รายละเอียด" in h or "detail" in h or "description" in h:
            if col_map["description"] == -1: col_map["description"] = idx
        elif "วันที่นัด" in h or "planned" in h or "นัดเข้า" in h:
            if col_map["planned_date"] == -1: col_map["planned_date"] = idx
        elif "เข้าจริง" in h or "actual" in h or "ทำจริง" in h or "เข้าทำ" in h:
            if col_map["actual_date"] == -1: col_map["actual_date"] = idx
        elif "deadline" in h or "deadlin" in h or "กำหนดเสร็จ" in h or "กำหนดส่ง" in h:
            if col_map["deadline_date"] == -1: col_map["deadline_date"] = idx
        elif "สถานที่" in h or "location" in h or "ศูนย์" in h:
            if col_map["location"] == -1: col_map["location"] = idx
        elif "สถานะ" in h or "status" in h:
            if col_map["status"] == -1: col_map["status"] = idx

    defaults = {
        "vehicle_no": 0,
        "job_title": 1,
        "description": 2,
        "planned_date": 3,
        "actual_date": 4,
        "deadline_date": 5,
        "location": 6,
        "status": 7
    }
    for key, val in defaults.items():
        if col_map[key] == -1:
            col_map[key] = val

    imported_count = 0
    errors = []

    for row_idx, row in enumerate(rows[1:], start=2):
        if not row or len(row) <= max(col_map.values()):
            errors.append(f"แถวที่ {row_idx}: จำนวนคอลัมน์ไม่ครบ")
            continue

        try:
            vehicle_no = str(row[col_map["vehicle_no"]]).strip()
            job_title = str(row[col_map["job_title"]]).strip()
            description = str(row[col_map["description"]]).strip() if col_map["description"] < len(row) else ""

            planned_date_raw = row[col_map["planned_date"]]
            planned_date = parse_date(planned_date_raw)
            actual_date = parse_date(row[col_map["actual_date"]]) if col_map["actual_date"] < len(row) else None
            deadline_date = parse_date(row[col_map["deadline_date"]]) if col_map["deadline_date"] < len(row) else None
            if not deadline_date:
                deadline_date = planned_date

            location = str(row[col_map["location"]]).strip() if col_map["location"] < len(row) else ""
            if not location:
                location = "ศูนย์ ISUZU (แหลมฉบัง อ.ศรีราชา)"

            status_val = str(row[col_map["status"]]).strip() if col_map["status"] < len(row) else "Planned"
            status_lower = status_val.lower()
            if "เสร็จ" in status_val or "complete" in status_lower:
                status_val = "Completed"
            elif "เนินการ" in status_val or "progress" in status_lower:
                status_val = "In Progress"
            elif "ยกเลิก" in status_val or "cancel" in status_lower:
                status_val = "Cancelled"
            elif "เลยกำหนด" in status_val or "overdue" in status_lower:
                status_val = "Overdue"
            else:
                status_val = "Planned"

            if not vehicle_no or not job_title or not planned_date:
                errors.append(f"แถวที่ {row_idx}: ทะเบียนรถ, ชื่องาน หรือวันที่นัดไม่ถูกต้อง")
                continue

            db_plan = PMPlan(
                vehicle_no=vehicle_no,
                job_title=job_title,
                description=description,
                planned_date=planned_date,
                actual_date=actual_date,
                deadline_date=deadline_date,
                location=location,
                status=status_val
            )

            enrich_plan_vehicle_fields(db_plan)
            if db_plan.status not in ["Completed", "Cancelled"] and db_plan.deadline_date < datetime.date.today():
                db_plan.status = "Overdue"

            db.add(db_plan)
            db.flush()
            add_pm_history(db, db_plan, "import", note=f"Import จากไฟล์ {filename}")
            imported_count += 1

        except Exception as e:
            errors.append(f"แถวที่ {row_idx}: เกิดข้อผิดพลาด {str(e)}")

    add_import_log(db, "plans", filename, "imported", total=max(len(rows)-1, 0), valid=imported_count, error=len(errors), imported=imported_count, errors=errors)
    db.commit()

    return {
        "success": True,
        "imported_count": imported_count,
        "errors": errors
    }



@app.get("/api/vehicles", response_model=List[VehicleMasterSchema])
def get_vehicles(db: Session = Depends(get_db)):
    return db.query(VehicleMaster).order_by(VehicleMaster.vehicle_no.asc()).all()

@app.post("/api/vehicles/sync-data-car")
def sync_vehicle_master(db: Session = Depends(get_db)):
    count = import_data_car_to_vehicle_master(db)
    add_import_log(db, "vehicles", "Data Car.csv", "imported", imported=count)
    db.commit()
    return {"success": True, "imported_or_updated": count}

@app.get("/api/plans/{plan_id}/history", response_model=List[PMHistorySchema])
def get_plan_history(plan_id: int, db: Session = Depends(get_db)):
    return db.query(PMHistory).filter(PMHistory.pm_plan_id == plan_id).order_by(PMHistory.changed_at.desc()).all()

@app.get("/api/import-logs", response_model=List[ImportLogSchema])
def get_import_logs(db: Session = Depends(get_db)):
    return db.query(ImportLog).order_by(ImportLog.created_at.desc()).limit(100).all()

@app.get("/api/notification-logs", response_model=List[NotificationLogSchema])
def get_notification_logs(db: Session = Depends(get_db)):
    return db.query(NotificationLog).order_by(NotificationLog.sent_at.desc()).limit(100).all()

@app.post("/api/plans/import/preview")
async def preview_plans_file(file: UploadFile = File(...), db: Session = Depends(get_db)):
    # Lightweight preview: reuses existing import parser by validating rows without inserting.
    content = await file.read()
    filename = (file.filename or "").lower()
    token = f"preview_{int(datetime.datetime.utcnow().timestamp())}"
    # Store uploaded file temporarily for confirm import.
    tmp_dir = os.path.join(os.path.dirname(__file__), "_import_tmp")
    os.makedirs(tmp_dir, exist_ok=True)
    tmp_path = os.path.join(tmp_dir, token + (".xlsx" if filename.endswith(".xlsx") else ".csv"))
    with open(tmp_path, "wb") as f:
        f.write(content)
    # Basic row count and column check for UX.
    total = 0; errors = []; valid = 0
    try:
        if filename.endswith(".xlsx"):
            wb = load_workbook(io.BytesIO(content), data_only=True); ws = wb.active
            rows = [[c if c is not None else "" for c in row] for row in ws.iter_rows(values_only=True) if any(str(v or '').strip() for v in row)]
        else:
            text = content.decode("utf-8-sig", errors="ignore") or content.decode("cp874", errors="ignore")
            rows = list(csv.reader(io.StringIO(text)))
        total = max(len(rows)-1, 0)
        if rows and len(rows[0]) < 2:
            errors.append("ไฟล์มีข้อมูลกองอยู่คอลัมน์เดียว กรุณาใช้ Export Excel .xlsx")
        else:
            valid = total
    except Exception as e:
        errors.append(str(e))
    add_import_log(db, "plans", filename, "previewed", total=total, valid=valid, error=len(errors), errors=errors)
    db.commit()
    return {"success": len(errors)==0, "token": token, "total_rows": total, "valid_rows": valid, "error_rows": len(errors), "errors": errors}

@app.post("/api/plans/import/confirm")
def confirm_plans_import(payload: ImportConfirmRequest, db: Session = Depends(get_db)):
    raise HTTPException(status_code=501, detail="Preview token พร้อมแล้ว แต่โหมด Confirm ผ่าน API แยกยังไม่เปิดใช้ใน UI นี้ ให้ใช้ปุ่มนำเข้า CSV/Excel เดิมเพื่อบันทึกจริง และระบบจะบันทึก Import_Log ให้อัตโนมัติ")


# ==================== v1.5 Productivity Edition ====================
class TaskActionRequest(BaseModel):
    note: Optional[str] = None
    reason: Optional[str] = None
    status: Optional[str] = None


def _task_state(db: Session, plan_id: int):
    row = db.query(PMTaskState).filter(PMTaskState.pm_plan_id == plan_id).first()
    if not row:
        row = PMTaskState(pm_plan_id=plan_id)
        db.add(row); db.flush()
    return row


def _priority(plan, today):
    if plan.status == "Overdue" or plan.deadline_date < today: return 5
    if plan.planned_date == today: return 4
    if plan.planned_date == today + datetime.timedelta(days=1): return 3
    if plan.planned_date <= today + datetime.timedelta(days=3): return 2
    return 1

@app.get("/api/assistant/today")
def assistant_today(db: Session = Depends(get_db)):
    today = datetime.date.today()
    plans = db.query(PMPlan).filter(
        PMPlan.status != "Cancelled",
        ((PMPlan.planned_date == today) | ((PMPlan.deadline_date < today) & (PMPlan.status != "Completed")))
    ).order_by(PMPlan.deadline_date.asc(), PMPlan.planned_date.asc()).all()
    states = {x.pm_plan_id:x for x in db.query(PMTaskState).filter(PMTaskState.pm_plan_id.in_([p.id for p in plans] or [-1])).all()}
    items=[]
    for plan in plans:
        st=states.get(plan.id)
        items.append({**_plan_to_dict(plan), "priority":_priority(plan,today), "paused":bool(st.paused) if st else False,
            "pause_reason":st.pause_reason if st else None, "followup_status":st.followup_status if st else None,
            "quick_note":st.quick_note if st else None})
    completed_today=db.query(PMPlan).filter(PMPlan.status=="Completed", PMPlan.actual_date==today).count()
    active=[x for x in items if x['status']!='Completed' and not x['paused']]
    return {"date":today.isoformat(), "total":len(items), "active":len(active), "completed":completed_today,
            "overdue":sum(1 for x in active if x['deadline_date'] < today.isoformat() or x['status']=='Overdue'),
            "paused":sum(1 for x in items if x['paused']), "items":items}

@app.post("/api/assistant/plans/{plan_id}/complete")
def one_click_complete(plan_id:int, payload:TaskActionRequest=TaskActionRequest(), db:Session=Depends(get_db)):
    plan=db.query(PMPlan).filter(PMPlan.id==plan_id).first()
    if not plan: raise HTTPException(404,"Plan not found")
    old=_plan_to_dict(plan)
    plan.status="Completed"; plan.actual_date=datetime.date.today(); plan.updated_at=datetime.datetime.utcnow(); plan.updated_by="one_click_complete"
    st=_task_state(db,plan_id); st.paused=False; st.completed_at=datetime.datetime.utcnow(); st.quick_note=payload.note or st.quick_note; st.updated_at=datetime.datetime.utcnow()
    db.add(PMHistory(pm_plan_id=plan.id,action="complete",changed_by="system",old_data=json.dumps(old,ensure_ascii=False,default=str),new_data=json.dumps(_plan_to_dict(plan),ensure_ascii=False,default=str),note=payload.note or "One Click Complete"))
    # v1.6: close matching Weekly PM Control items automatically.
    vehicle_key = extract_car_no(plan.vehicle_no)
    weekly_rows = db.query(WeeklyCampaignItem).filter(WeeklyCampaignItem.vehicle_no == vehicle_key, WeeklyCampaignItem.status != "completed").all()
    for weekly in weekly_rows:
        weekly.status = "completed"
        weekly.actual_date = datetime.date.today()
        weekly.pm_plan_id = plan.id
        weekly.updated_at = datetime.datetime.utcnow()
    db.commit()
    return {"success":True,"message":"ปิดงานเรียบร้อย","plan":_plan_to_dict(plan)}

@app.post("/api/assistant/plans/{plan_id}/pause")
def pause_task(plan_id:int,payload:TaskActionRequest,db:Session=Depends(get_db)):
    plan=db.query(PMPlan).filter(PMPlan.id==plan_id).first()
    if not plan: raise HTTPException(404,"Plan not found")
    st=_task_state(db,plan_id); st.paused=True; st.pause_reason=payload.reason or "พักการติดตาม"; st.updated_at=datetime.datetime.utcnow()
    db.add(PMHistory(pm_plan_id=plan.id,action="pause",changed_by="system",note=st.pause_reason)); db.commit()
    return {"success":True}

@app.post("/api/assistant/plans/{plan_id}/resume")
def resume_task(plan_id:int,db:Session=Depends(get_db)):
    st=_task_state(db,plan_id); st.paused=False; st.pause_reason=None; st.updated_at=datetime.datetime.utcnow(); db.commit(); return {"success":True}

@app.post("/api/assistant/plans/{plan_id}/followup")
def followup_task(plan_id:int,payload:TaskActionRequest,db:Session=Depends(get_db)):
    plan=db.query(PMPlan).filter(PMPlan.id==plan_id).first()
    if not plan: raise HTTPException(404,"Plan not found")
    st=_task_state(db,plan_id); st.followup_status=payload.status or "โทรแล้ว"; st.quick_note=payload.note or st.quick_note; st.updated_at=datetime.datetime.utcnow()
    db.add(PMHistory(pm_plan_id=plan.id,action="followup",changed_by="system",note=f"{st.followup_status}: {st.quick_note or ''}")); db.commit()
    return {"success":True}

@app.get("/api/assistant/weekly-summary")
def weekly_summary(db:Session=Depends(get_db)):
    today=datetime.date.today(); start=today-datetime.timedelta(days=today.weekday()); end=start+datetime.timedelta(days=6)
    rows=db.query(PMPlan).filter(PMPlan.planned_date>=start,PMPlan.planned_date<=end,PMPlan.status!="Cancelled").all()
    done=[r for r in rows if r.status=="Completed"]
    overdue=[r for r in rows if r.status!="Completed" and r.deadline_date<today]
    by_job={}; by_loc={}
    for r in rows: by_job[r.job_title]=by_job.get(r.job_title,0)+1; by_loc[r.location]=by_loc.get(r.location,0)+1
    return {"start":start.isoformat(),"end":end.isoformat(),"total":len(rows),"completed":len(done),"pending":len(rows)-len(done),"overdue":len(overdue),"completion_rate":round(len(done)*100/len(rows),1) if rows else 0,
            "top_job":max(by_job,key=by_job.get) if by_job else "-","top_location":max(by_loc,key=by_loc.get) if by_loc else "-"}

def _assistant_send(db, text, kind):
    token=_get_setting_value(db,"line_channel_access_token",""); target=_get_setting_value(db,"line_target_id","")
    if not token or not target: return False
    result=send_line_message_debug(token,target,text)
    db.add(NotificationLog(channel="LINE",target_id=target,message=text,status="success" if result.get("success") else "failed",response=json.dumps({"kind":kind,"debug":result},ensure_ascii=False,default=str))); db.commit()
    return bool(result.get("success"))

def morning_brief_job():
    db=database.SessionLocal()
    try:
        today=datetime.date.today(); rows=db.query(PMPlan).filter(PMPlan.status!="Cancelled", PMPlan.status!="Completed", ((PMPlan.planned_date==today)|(PMPlan.deadline_date<today))).all()
        overdue=sum(1 for r in rows if r.deadline_date<today)
        text=f"🌅 Morning Brief {today.strftime('%d/%m/%Y')}\n📋 งานที่ต้องดูวันนี้ {len(rows)} คัน\n🔴 Overdue {overdue} คัน\n👉 เปิด My Today เพื่อจัดการงาน"
        _assistant_send(db,text,"morning_brief")
    finally: db.close()

def followup_job():
    db=database.SessionLocal()
    try:
        today=datetime.date.today(); paused={x.pm_plan_id for x in db.query(PMTaskState).filter(PMTaskState.paused==True).all()}
        rows=db.query(PMPlan).filter(PMPlan.status!="Completed",PMPlan.status!="Cancelled",PMPlan.planned_date<=today).all(); rows=[r for r in rows if r.id not in paused]
        if rows:
            lines=[f"• {r.vehicle_no} | {r.job_title} | {r.location}" for r in rows[:10]]
            _assistant_send(db,"⚠️ งานที่ต้องติดตาม\n"+"\n".join(lines)+(f"\nและอีก {len(rows)-10} คัน" if len(rows)>10 else ""),"followup")
    finally: db.close()

def evening_summary_job():
    db=database.SessionLocal()
    try:
        today=datetime.date.today(); done=db.query(PMPlan).filter(PMPlan.status=="Completed",PMPlan.actual_date==today).count(); pending=db.query(PMPlan).filter(PMPlan.status!="Completed",PMPlan.status!="Cancelled",PMPlan.planned_date<=today).count()
        _assistant_send(db,f"🌙 สรุปงานวันนี้ {today.strftime('%d/%m/%Y')}\n✅ เสร็จ {done} คัน\n⏳ ค้างติดตาม {pending} คัน","evening_summary")
    finally: db.close()

def weekly_summary_job():
    db=database.SessionLocal()
    try:
        today=datetime.date.today(); start=today-datetime.timedelta(days=today.weekday()); rows=db.query(PMPlan).filter(PMPlan.planned_date>=start,PMPlan.planned_date<=today,PMPlan.status!="Cancelled").all(); done=sum(1 for r in rows if r.status=="Completed"); overdue=sum(1 for r in rows if r.status!="Completed" and r.deadline_date<today)
        _assistant_send(db,f"📊 Weekly Summary\nแผนทั้งหมด {len(rows)} คัน\n✅ Completed {done} คัน\n🔴 Overdue {overdue} คัน\nอัตราปิดงาน {round(done*100/len(rows),1) if rows else 0}%","weekly_summary")
    finally: db.close()

def update_assistant_scheduler(db:Session):
    for jid in ["v15_morning","v15_followup","v15_evening","v15_weekly"]:
        if scheduler.get_job(jid): scheduler.remove_job(jid)
    if _get_setting_value(db,"assistant_notifications_enabled","false").lower()!="true": return
    def hm(key,default):
        try:return map(int,_get_setting_value(db,key,default).split(':'))
        except:return map(int,default.split(':'))
    h,m=hm("morning_brief_time","08:00"); scheduler.add_job(morning_brief_job,CronTrigger(hour=h,minute=m,timezone="Asia/Bangkok"),id="v15_morning",replace_existing=True)
    h,m=hm("followup_time","14:00"); scheduler.add_job(followup_job,CronTrigger(hour=h,minute=m,timezone="Asia/Bangkok"),id="v15_followup",replace_existing=True)
    h,m=hm("evening_summary_time","17:30"); scheduler.add_job(evening_summary_job,CronTrigger(hour=h,minute=m,timezone="Asia/Bangkok"),id="v15_evening",replace_existing=True)
    h,m=hm("weekly_summary_time","17:00"); scheduler.add_job(weekly_summary_job,CronTrigger(day_of_week=_get_setting_value(db,"weekly_summary_day","fri"),hour=h,minute=m,timezone="Asia/Bangkok"),id="v15_weekly",replace_existing=True)

@app.post("/api/assistant/scheduler/enable")
def assistant_scheduler_enable(enabled:bool=Query(...),db:Session=Depends(get_db)):
    row=db.query(Setting).filter(Setting.key=="assistant_notifications_enabled").first(); row.value=str(enabled).lower(); db.commit(); update_assistant_scheduler(db); return {"enabled":enabled}

# Initialize v1.5 assistant scheduler
_v15db=database.SessionLocal()
try: update_assistant_scheduler(_v15db)
finally: _v15db.close()



def _build_v161_summary(db: Session, period: str = "now"):
    today=datetime.date.today()
    rows=db.query(PMPlan).filter(PMPlan.status!="Cancelled").all()
    today_rows=[r for r in rows if r.planned_date==today]
    overdue=[r for r in rows if r.status!="Completed" and r.deadline_date and r.deadline_date<today]
    completed=[r for r in rows if r.status=="Completed" and r.actual_date==today]
    pending=[r for r in today_rows if r.status!="Completed"]
    title={"morning":"🌅 สรุปงาน PM ช่วงเช้า","noon":"☀️ สรุปงาน PM ช่วงกลางวัน","evening":"🌙 สรุปงาน PM ช่วงเย็น"}.get(period,"📋 สรุปงาน PM")
    lines=[title, f"วันที่ {today.strftime('%d/%m/%Y')}", "", f"📌 วันนี้ทั้งหมด {len(today_rows)} คัน", f"✅ เสร็จแล้ว {len(completed)} คัน", f"⏳ ยังไม่เสร็จ {len(pending)} คัน", f"🔴 Overdue {len(overdue)} คัน"]
    if pending:
        lines += ["", "รายการที่ยังต้องติดตาม:"] + [f"• {r.vehicle_no} | {r.job_title} | {r.location}" for r in pending[:12]]
        if len(pending)>12: lines.append(f"และอีก {len(pending)-12} คัน")
    return "\n".join(lines)

@app.get("/api/reports/preview")
def preview_report(period:str="now", db:Session=Depends(get_db)):
    return {"period":period,"message":_build_v161_summary(db,period)}

@app.post("/api/reports/send")
def send_report(period:str="now", db:Session=Depends(get_db)):
    text=_build_v161_summary(db,period)
    ok=_assistant_send(db,text,f"manual_{period}")
    if not ok: raise HTTPException(status_code=500,detail="ส่งสรุปรายงาน LINE ไม่สำเร็จ")
    return {"success":True,"message":text}

def _v161_report_job(period:str):
    db=database.SessionLocal()
    try:
        text=_build_v161_summary(db,period)
        _assistant_send(db,text,f"report_{period}")
    finally: db.close()

def update_v161_report_scheduler(db:Session):
    for jid in ["v161_report_morning","v161_report_noon","v161_report_evening"]:
        if scheduler.get_job(jid): scheduler.remove_job(jid)
    def add_slot(period,key_time,key_enabled,default):
        if _get_setting_value(db,key_enabled,"true").lower()!="true": return
        try: h,m=map(int,_get_setting_value(db,key_time,default).split(':'))
        except: h,m=map(int,default.split(':'))
        scheduler.add_job(lambda p=period:_v161_report_job(p),CronTrigger(hour=h,minute=m,timezone="Asia/Bangkok"),id=f"v161_report_{period}",replace_existing=True)
    add_slot("morning","report_morning_time","report_morning_enabled","07:30")
    add_slot("noon","report_noon_time","report_noon_enabled","12:00")
    add_slot("evening","report_evening_time","report_evening_enabled","17:00")

@app.on_event("startup")
def _initialize_v161_scheduler():
    db=database.SessionLocal()
    try: update_v161_report_scheduler(db)
    finally: db.close()

# Serve Frontend static files
try:
    app.mount("/static", StaticFiles(directory="static"), name="static")
except Exception as e:
    print(f"[Warning] Failed to mount static files: {e}")



# ==================== v1.6 Weekly PM Control ====================
class WeeklyItemUpdate(BaseModel):
    status: str
    note: Optional[str] = None


def _weekly_item_dict(x):
    return {"id":x.id,"campaign_id":x.campaign_id,"vehicle_no":x.vehicle_no,"registration":x.registration,
            "pm_group":x.pm_group,"week_no":x.week_no,"lot_name":x.lot_name,"status":x.status,
            "actual_date":x.actual_date.isoformat() if x.actual_date else None,"note":x.note,"pm_plan_id":x.pm_plan_id}

@app.get("/api/weekly-control")
def weekly_control(year: Optional[int]=None, month: Optional[int]=None, week: Optional[int]=None, item_status: Optional[str]=None, db:Session=Depends(get_db)):
    today=datetime.date.today(); year=year or today.year; month=month or today.month
    campaigns=db.query(WeeklyCampaign).filter(WeeklyCampaign.year==year,WeeklyCampaign.month==month).order_by(WeeklyCampaign.id.desc()).all()
    ids=[c.id for c in campaigns]
    q=db.query(WeeklyCampaignItem).filter(WeeklyCampaignItem.campaign_id.in_(ids or [-1]))
    if week: q=q.filter(WeeklyCampaignItem.week_no==week)
    if item_status and item_status!="all": q=q.filter(WeeklyCampaignItem.status==item_status)
    rows=q.order_by(WeeklyCampaignItem.week_no,WeeklyCampaignItem.vehicle_no).all()
    all_rows=db.query(WeeklyCampaignItem).filter(WeeklyCampaignItem.campaign_id.in_(ids or [-1])).all()
    weeks={}
    for w in range(1,6):
        wr=[r for r in all_rows if r.week_no==w]; done=sum(1 for r in wr if r.status=="completed")
        weeks[str(w)]={"total":len(wr),"completed":done,"pending":sum(1 for r in wr if r.status=="pending"),"deferred":sum(1 for r in wr if r.status=="deferred"),"rate":round(done*100/len(wr),1) if wr else 0}
    return {"year":year,"month":month,"campaigns":[{"id":c.id,"name":c.name,"note":c.note} for c in campaigns],
            "summary":{"total":len(all_rows),"completed":sum(1 for r in all_rows if r.status=="completed"),"pending":sum(1 for r in all_rows if r.status=="pending"),"deferred":sum(1 for r in all_rows if r.status=="deferred")},
            "weeks":weeks,"items":[_weekly_item_dict(r) for r in rows]}

@app.patch("/api/weekly-control/items/{item_id}")
def update_weekly_item(item_id:int,payload:WeeklyItemUpdate,db:Session=Depends(get_db)):
    row=db.query(WeeklyCampaignItem).filter(WeeklyCampaignItem.id==item_id).first()
    if not row: raise HTTPException(404,"Weekly item not found")
    if payload.status not in {"pending","completed","deferred"}: raise HTTPException(400,"Invalid status")
    row.status=payload.status; row.note=payload.note if payload.note is not None else row.note; row.updated_at=datetime.datetime.utcnow()
    row.actual_date=datetime.date.today() if payload.status=="completed" else None
    db.commit(); return {"success":True,"item":_weekly_item_dict(row)}

@app.delete("/api/weekly-control/items/{item_id}")
def delete_weekly_item(item_id:int,db:Session=Depends(get_db)):
    row=db.query(WeeklyCampaignItem).filter(WeeklyCampaignItem.id==item_id).first()
    if not row: raise HTTPException(404,"Weekly item not found")
    db.delete(row); db.commit(); return {"success":True}

@app.post("/api/weekly-control/import")
async def import_weekly_control(file:UploadFile=File(...),year:int=Query(...),month:int=Query(...),campaign_name:str=Query("Weekly PM Control"),db:Session=Depends(get_db)):
    raw=await file.read(); name=(file.filename or "").lower(); rows=[]
    if name.endswith('.xlsx'):
        wb=load_workbook(io.BytesIO(raw),data_only=True); ws=wb.active
        values=list(ws.iter_rows(values_only=True)); headers=[str(x or '').strip() for x in (values[0] if values else [])]; rows=[list(r) for r in values[1:]]
    else:
        text_data=None
        for enc in ('utf-8-sig','cp874','tis-620'):
            try: text_data=raw.decode(enc); break
            except: pass
        if text_data is None: raise HTTPException(400,'อ่านไฟล์ไม่ได้')
        sample=text_data[:3000]; delim=',' if sample.count(',')>=sample.count(';') else ';'
        values=list(csv.reader(io.StringIO(text_data),delimiter=delim)); headers=[str(x or '').strip() for x in (values[0] if values else [])]; rows=values[1:]
    norm=lambda x: ''.join(str(x or '').lower().replace('_',' ').split())
    hm={norm(h):i for i,h in enumerate(headers)}
    def idx(*keys):
        for k in keys:
            nk=norm(k)
            if nk in hm:return hm[nk]
        return None
    i_vehicle=idx('เบอร์รถ','รถเบอร์','vehicle no','vehicle_no','เลขรถ'); i_reg=idx('ทะเบียนรถ','ทะเบียน','registration')
    i_group=idx('มย.','กลุ่ม','pm group','group'); i_week=idx('week','สัปดาห์','สัปดาห์ที่'); i_lot=idx('lot','wave','ล็อต')
    if i_vehicle is None and i_reg is None: raise HTTPException(400,'ไม่พบคอลัมน์ เบอร์รถ หรือ ทะเบียนรถ')
    campaign=WeeklyCampaign(name=campaign_name or f'Weekly PM {month}/{year}',year=year,month=month); db.add(campaign); db.flush()
    imported=0; skipped=0
    for r in rows:
        def val(i): return str(r[i] or '').strip() if i is not None and i<len(r) else ''
        vehicle=val(i_vehicle) or val(i_reg)
        if not vehicle: skipped+=1; continue
        try: week_no=int(float(val(i_week))) if val(i_week) else 1
        except: week_no=1
        week_no=max(1,min(5,week_no))
        db.add(WeeklyCampaignItem(campaign_id=campaign.id,vehicle_no=vehicle,registration=val(i_reg),pm_group=val(i_group),week_no=week_no,lot_name=val(i_lot),status='pending'))
        imported+=1
    db.add(ImportLog(import_type='weekly_control',filename=file.filename,status='imported',total_rows=len(rows),valid_rows=imported,error_rows=skipped,imported_rows=imported,errors=f'Skipped {skipped} rows' if skipped else None))
    db.commit(); return {"success":True,"campaign_id":campaign.id,"imported":imported,"skipped":skipped}

@app.get("/api/weekly-control/template")
def weekly_template():
    wb=Workbook(); ws=wb.active; ws.title='Weekly PM Control'
    ws.append(['เบอร์รถ','ทะเบียนรถ','มย.','Week','Lot','หมายเหตุ'])
    ws.append(['245','79-8760','A1',1,'Lot A',''])
    ws.freeze_panes='A2'
    out=io.BytesIO(); wb.save(out); out.seek(0)
    return StreamingResponse(out,media_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',headers={'Content-Disposition':'attachment; filename=weekly_pm_control_template.xlsx'})

@app.post("/api/weekly-control/send-line")
def send_weekly_line(year:int=Query(...),month:int=Query(...),week:int=Query(...),db:Session=Depends(get_db)):
    campaigns=db.query(WeeklyCampaign).filter(WeeklyCampaign.year==year,WeeklyCampaign.month==month).all(); ids=[c.id for c in campaigns]
    rows=db.query(WeeklyCampaignItem).filter(WeeklyCampaignItem.campaign_id.in_(ids or [-1]),WeeklyCampaignItem.week_no==week,WeeklyCampaignItem.status=='pending').order_by(WeeklyCampaignItem.vehicle_no).all()
    title=f'📣 ติดตามแผน PM เดือน {month}/{year} | สัปดาห์ที่ {week}'
    body='\n'.join([f'{i+1}. รถ {r.vehicle_no}' + (f' | กลุ่ม {r.pm_group}' if r.pm_group else '') for i,r in enumerate(rows[:40])]) or '- ไม่มีรถค้างติดตาม'
    text=f'{title}\nเหลือค้าง {len(rows)} คัน\n====================\n{body}'
    ok=_assistant_send(db,text,'weekly_control')
    return {"success":ok,"message":text,"pending":len(rows)}


@app.get("/")
def read_root():
    return FileResponse("static/index.html")
